import openpyxl
import json
import os

def extract_data():
    excel_path = os.path.join(os.path.dirname(__file__), '..', 'errorr-seo-master-single-sheet.xlsx')
    excel_path = os.path.abspath(excel_path)
    
    print(f"Loading Excel file from: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb['errorr.in SEO Master']

    # Headers in row 258
    headers = []
    for c in range(1, 26):
        val = sheet.cell(row=258, column=c).value
        headers.append(str(val).strip() if val else f"col_{c}")

    pages = []
    page_index = {}

    for r in range(259, 1259):
        page_id = sheet.cell(row=r, column=1).value
        if not page_id:
            continue
        
        row_data = {}
        for idx, h in enumerate(headers):
            val = sheet.cell(row=r, column=idx + 1).value
            row_data[h] = val

        url_slug = str(row_data.get('URL Slug', '') or '').strip()
        if not url_slug.startswith('/'):
            url_slug = '/' + url_slug
        if not url_slug.endswith('/'):
            url_slug = url_slug + '/'
        
        # Standardize slug
        row_data['clean_slug'] = url_slug

        # Split H2 outline
        h2_raw = row_data.get('H2 Outline (pipe-separated)', '') or ''
        row_data['h2_list'] = [h.strip() for h in str(h2_raw).split('|') if h.strip()]

        pages.append(row_data)
        page_index[url_slug] = row_data

    out_dir = os.path.join(os.path.dirname(__file__), '..', 'src', 'data')
    os.makedirs(out_dir, exist_ok=True)

    json_path = os.path.join(out_dir, 'pageMaster.json')
    index_path = os.path.join(out_dir, 'pageMasterIndex.json')

    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(pages, f, indent=2, ensure_ascii=False)

    with open(index_path, 'w', encoding='utf-8') as f:
        json.dump(page_index, f, indent=2, ensure_ascii=False)

    print(f"Successfully extracted {len(pages)} pages.")
    print(f"Saved to {json_path} and {index_path}")

if __name__ == '__main__':
    extract_data()
